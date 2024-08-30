from openpyxl.workbook.workbook import Worksheet
from django.http import HttpResponse
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


# Create your views here.
def setHeader(sheet: Worksheet, names, font=Font(size=12, bold=True),
              fill1=PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
              fill2=PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')):
    workbook = Workbook()
    s = workbook.active
    for index, name in enumerate(names):
        sheet.cell(1, index + 1, name)
    turn = False
    for cell in sheet[1]:
        cell.font = font
        if turn:
            cell.fill = fill1
        else:
            cell.fill = fill2
        turn = not turn


def writeData(sheet: Worksheet, attrs, datas):
    for row_index, row_data in enumerate(datas, start=2):
        for col_index, attr in enumerate(attrs, start=1):
            sheet.cell(row_index, col_index, row_data[attr])


def exportData(names, attrs, datas):
    wb = Workbook()
    sheet = wb.active
    setHeader(sheet, names)
    writeData(sheet, attrs, datas)
    return wb


def exportStudents(request):
    # if request.method == 'OPTIONS':
    #     # 处理 OPTIONS 请求，返回允许的方法和头部
    #     response = JsonResponse({'message': 'Allowed'}, status=200)
    #     response['Access-Control-Allow-Origin'] = '*'  # 允许所有来源
    #     response['Access-Control-Allow-Methods'] = 'POST, GET, OPTIONS'  # 允许的方法
    #     response['Access-Control-Allow-Headers'] = 'Content-Type'  # 允许的头部
    #     return response
    datas = json.loads(request.body)['students']
    names = ['姓名', '性别', '民族', '邮箱', '电话号码', '毕业学校', '毕业专业', '入学日期', '学号', '生日',
             '当前毕业类型', '状态']
    attrs = ['name', 'gender', 'ethnicity', 'email', 'phone', 'university', 'department', 'enrollment_date',
             'student_id', 'birthdate', 'graduate_type', 'status']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response


def exportTeachers(request):
    datas = json.loads(request.body)['teachers']
    names = ['姓名', '性别', '民族', '邮箱', '电话号码', '职称', '毕业学校', '毕业专业', '入职日期', '教师工号', '生日',
             '办公地点',
             '状态']
    attrs = ['name', 'gender', 'ethnicity', 'email', 'phone', 'professional_title', 'university', 'department',
             'enrollment_date', 'teacher_id', 'birthdate', 'office_address', 'status']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response


def exportAwards(request):
    datas = json.loads(request.body)['awards']
    names = ['名称', '级别', '颁奖单位', '获奖日期', '获奖者', '原因']
    attrs = ['title', 'level', 'awarding_unit', 'presentation_date', 'prizewinner', 'reason']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response


def exportMonographs(request):
    datas = json.loads(request.body)['monographs']
    names = ['名称', '作者', '发表日期', '出版商', '关键字', '语言', 'ISBN', '概要']
    attrs = ['title', 'author', 'publication_date', 'publisher', 'keyword', 'language', 'isbn', 'abstract']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response


def exportConferencePapers(request):
    datas = json.loads(request.body)['conferencePapers']
    names = ['名称', '作者', '主题', '会议', '发表日期', '关键字', '语言', 'DOI', '概要', '引用数', '下载数',
             '参考文献']
    attrs = ['title', 'author', 'subject', 'conference', 'publication_date', 'keyword', 'language', 'doi', 'abstract',
             'citation_number', 'download_count', 'references']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response


def exportJournalPapers(request):
    datas = json.loads(request.body)['journalPapers']
    names = ['名称', '作者', '主题', '期刊', '发表日期', '关键字', '语言', 'DOI', '概要', '引用数', '下载数',
             '参考文献']
    attrs = ['title', 'author', 'subject', 'journal', 'publication_date', 'keyword', 'language', 'doi', 'abstract',
             'citation_number', 'download_count', 'references']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response


def exportPatents(request):
    datas = json.loads(request.body)['patents']
    names = ['名称', '发明者', '申请者', '持有人', '申请日期', '授权日期', '专利号', '领域', '国家', '状态', '概要']
    attrs = ['title', 'inventor', 'applicant', 'right_holder', 'application_date', 'authorization_date',
             'patent_number', 'domain', 'country', 'legal_status', 'abstract']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response

def exportProjects(request):
    datas = json.loads(request.body)['projects']
    names = ['名称', '类型', '负责人', '核心成员', '参与人员', '项目号', '开始日期', '结束日期', '经费', '状态']
    attrs = ['title', 'type', 'leader', 'core_member', 'participant', 'project_number',
             'start_date', 'end_date', 'fund', 'status']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response

def exportSoftwareCopyrights(request):
    datas = json.loads(request.body)['softwareCopyrights']
    names = ['名称', '类型', '持有者', '注册号', '注册日期', '开发语言', '有效期截至', '版本']
    attrs = ['title', 'type', 'owner', 'registration_number', 'registration_date', 'development_language',
             'valid_period_deadline', 'version']
    wb = exportData(names, attrs, datas)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    wb.save(response)
    return response